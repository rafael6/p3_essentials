#pip install openpyxl
import openpyxl
import os

os.chdir('c:\\users\\al\\documents')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

workbook.get_sheet_names()

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

cell = sheet['A1']
print(cell.value)

print(str(sheet['A1'].value))

print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=1, column=2).value)



